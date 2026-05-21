"""
使用pandoc将多种文档格式转换为markdown，提取图片和嵌入附件

支持格式：
- docx: Word文档，提取图片和OLE附件
- xlsx/xls: Excel文件，转换为markdown表格
"""
import subprocess
import os
import re
import shutil
import zipfile
import struct
import io
import tempfile
from pathlib import Path
from PIL import Image
from openpyxl import load_workbook
import pandas as pd


def get_ole10native_filename(data):
    """从Package对象获取原始文件名和嵌入数据"""
    if data[:8] != b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1':
        return None, None

    sector_size = 512
    dir_start = struct.unpack('<I', data[48:52])[0]
    dir_offset = 512 + dir_start * sector_size

    for i in range(16):
        entry_offset = dir_offset + i * 128
        if entry_offset + 128 > len(data):
            break

        entry_data = data[entry_offset:entry_offset + 128]
        name_len = struct.unpack('<H', entry_data[64:66])[0]
        if name_len > 0 and name_len <= 64:
            name_bytes = entry_data[:name_len]
            try:
                name = name_bytes.decode('utf-16-le', errors='replace').rstrip('\x00')
            except:
                continue

            if name == '\x01Ole10Native':
                start_sector = struct.unpack('<I', entry_data[116:120])[0]
                stream_size = struct.unpack('<I', entry_data[120:124])[0]

                stream_offset = 512 + start_sector * sector_size
                if stream_offset < len(data):
                    stream_data = data[stream_offset:stream_offset + stream_size]

                    # Package Ole10Native格式:
                    # [4 bytes] 数据大小
                    # [2 bytes] flags
                    # [GBK字符串] 原始文件名，null结尾
                    # [ANSI字符串] 临时文件路径，null结尾
                    # [额外元数据] 可能包含另一个路径
                    # [文件数据] 实际嵌入内容

                    # 从位置6开始找文件名null结束
                    filename_start = 6
                    filename_end = stream_data.find(b'\x00', filename_start)
                    if filename_end != -1 and filename_end < 300:
                        filename_bytes = stream_data[filename_start:filename_end]
                        filename = filename_bytes.decode('gbk', errors='replace')

                        # 找临时路径结束
                        temp_path_end = stream_data.find(b'\x00', filename_end + 1)

                        # 在数据中搜索文件类型签名，获取实际内容
                        # PNG签名
                        png_pos = stream_data.find(b'\x89PNG\r\n\x1a\n', temp_path_end)
                        if png_pos != -1:
                            embedded_data = stream_data[png_pos:]
                            # 找PNG结束(IEND)
                            iend_pos = embedded_data.find(b'IEND')
                            if iend_pos != -1:
                                embedded_data = embedded_data[:iend_pos + 12]
                            return filename, embedded_data

                        # BMP签名
                        bmp_pos = stream_data.find(b'BM', temp_path_end)
                        if bmp_pos != -1:
                            # BMP数据需要从BM开始到文件结束
                            # BMP文件头包含文件大小信息
                            bmp_header = stream_data[bmp_pos:bmp_pos + 54]
                            if len(bmp_header) >= 54:
                                bmp_file_size = struct.unpack('<I', bmp_header[2:6])[0]
                                embedded_data = stream_data[bmp_pos:bmp_pos + bmp_file_size]
                                return filename, embedded_data

                        # JPEG签名
                        jpeg_pos = stream_data.find(b'\xff\xd8\xff', temp_path_end)
                        if jpeg_pos != -1:
                            # 找JPEG结束标记
                            jpeg_end = stream_data.find(b'\xff\xd9', jpeg_pos)
                            if jpeg_end != -1:
                                embedded_data = stream_data[jpeg_pos:jpeg_end + 2]
                                return filename, embedded_data

                        # 如果没找到已知签名，从temp_path_end+1开始作为数据
                        embedded_data = stream_data[temp_path_end + 1:]
                        return filename, embedded_data
    return None, None


def extract_xlsx_from_ole(data):
    """从Excel OLE对象提取xlsx - 改进版本"""
    # xlsx是zip格式，必须找到完整的zip结构
    # 在OLE对象中，xlsx数据通常从特定位置开始

    # 找到第一个PK (local file header) - 通常在固定偏移后
    pk_local = data.find(b'PK\x03\x04')
    if pk_local == -1:
        return None

    # 找到EOCD (End of Central Directory)
    pk_eocd = data.find(b'PK\x05\x06')
    if pk_eocd == -1:
        return None

    # 读取EOCD获取注释长度
    eocd_data = data[pk_eocd:pk_eocd + 22]
    comment_len = struct.unpack('<H', eocd_data[20:22])[0]
    eocd_end = pk_eocd + 22 + comment_len

    # 提取完整zip数据
    xlsx_data = data[pk_local:eocd_end]

    # 验证zip完整性
    try:
        z = zipfile.ZipFile(io.BytesIO(xlsx_data))
        # 检查关键文件是否存在
        if '[Content_Types].xml' not in z.namelist():
            return None
        return xlsx_data
    except:
        return None


def extract_ole_objects(docx_path, assets_dir, base_name):
    """从docx中提取OLE嵌入对象（按需创建目录）"""
    ole_info = {}  # image_name -> {'original_name': str, 'saved_name': str, 'is_image': bool, 'image_idx': int}
    image_idx = 1  # 图片编号计数器

    # 目录路径（按需创建）
    images_dir = assets_dir / "images"
    attachments_dir = assets_dir / "attachments"

    with zipfile.ZipFile(docx_path, 'r') as z:
        # 读取关系文件
        rels_xml = z.read('word/_rels/document.xml.rels').decode('utf-8')

        # 解析所有关系
        rels = {}
        for match in re.finditer(r'<Relationship Id="rId(\d+)"[^>]*Type="[^"]*relationships/([^"]+)"[^>]*Target="([^"]+)"', rels_xml):
            rId = int(match.group(1))
            rel_type = match.group(2)
            target = match.group(3)
            rels[rId] = (rel_type, target)

        # 找OLE对象和对应图片的映射
        ole_to_image = {}
        for rId, (rel_type, target) in sorted(rels.items()):
            if rel_type == 'oleObject' and 'embeddings/' in target:
                ole_num_match = re.search(r'oleObject(\d+)', target)
                if ole_num_match:
                    ole_num = int(ole_num_match.group(1))
                    next_rId = rId + 1
                    if next_rId in rels:
                        next_type, next_target = rels[next_rId]
                        if next_type == 'image' and 'media/' in next_target:
                            image_name = next_target.split('/')[-1]
                            ole_to_image[ole_num] = image_name

        # 读取document.xml获取OLE对象的ProgID
        doc_xml = z.read('word/document.xml').decode('utf-8')

        ole_progids = {}
        for match in re.finditer(r'<o:OLEObject[^>]*ProgID="([^"]+)"[^>]*r:id="rId(\d+)"', doc_xml):
            prog_id = match.group(1)
            ole_rId = int(match.group(2))
            ole_progids[ole_rId] = prog_id

        # 处理每个OLE对象
        for ole_rId, prog_id in ole_progids.items():
            if ole_rId not in rels:
                continue
            _, ole_target = rels[ole_rId]
            ole_num_match = re.search(r'oleObject(\d+)', ole_target)
            if not ole_num_match:
                continue
            ole_num = int(ole_num_match.group(1))

            ole_path = f'word/{ole_target}'
            if ole_path not in z.namelist():
                continue

            ole_data = z.read(ole_path)
            image_name = ole_to_image.get(ole_num)

            original_name = None
            embedded_data = None
            is_image = False

            if prog_id == 'Package':
                original_name, embedded_data = get_ole10native_filename(ole_data)
                # 检查是否是图片文件
                if original_name and original_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp')):
                    is_image = True
                else:
                    is_image = False
            elif prog_id == 'Excel.Sheet.12':
                xlsx_data = extract_xlsx_from_ole(ole_data)
                if xlsx_data:
                    try:
                        # 用openpyxl读取数据
                        wb_source = load_workbook(io.BytesIO(xlsx_data), data_only=True)

                        # 获取工作表名称作为文件名
                        sheet_names = wb_source.sheetnames
                        if sheet_names:
                            original_name = sheet_names[0] + '.xlsx'
                        else:
                            original_name = f'附件{ole_num}.xlsx'

                        # 用pandas+xlsxwriter重新生成xlsx文件
                        new_buffer = io.BytesIO()
                        with pd.ExcelWriter(new_buffer, engine='xlsxwriter') as writer:
                            for sheet_name in sheet_names:
                                ws = wb_source[sheet_name]
                                data = []
                                for row in ws.iter_rows(values_only=True):
                                    data.append(list(row))
                                df = pd.DataFrame(data)
                                df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)

                        embedded_data = new_buffer.getvalue()
                        is_image = False

                        # 同时生成md格式的文本文件
                        md_content = ""
                        for sheet_name in sheet_names:
                            ws = wb_source[sheet_name]
                            md_content += f"## {sheet_name}\n\n"

                            # 读取数据转为表格格式
                            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                                # 过滤空行
                                if any(cell is not None for cell in row):
                                    cells = [str(cell) if cell is not None else '' for cell in row]
                                    md_content += "| " + " | ".join(cells) + " |\n"
                                    # 第一行添加表头分隔线
                                    if row_idx == 0:
                                        md_content += "| " + " | ".join(["---"] * len(cells)) + " |\n"
                            md_content += "\n"

                        wb_source.close()

                        # md文件名（与xlsx同名，扩展名改为.md）
                        md_name = original_name.replace('.xlsx', '.md')

                        print(f"Excel重新生成(xlsxwriter): {original_name}")
                        print(f"同时生成MD文本: {md_name}")

                        # 保存md文件（稍后在统一位置处理）
                        # 这里先记录md内容，在后面保存xlsx时同时保存md
                        ole_info[image_name] = {
                            'xlsx_name': original_name,
                            'xlsx_data': embedded_data,
                            'md_name': md_name,
                            'md_content': md_content,
                            'is_image': False
                        }
                        # 跳过后续的统一保存逻辑
                        continue
                    except Exception as e:
                        print(f"Excel处理失败(ole{ole_num}): {e}")
                        continue
            elif prog_id == 'Visio.Drawing.11':
                embedded_data = ole_data
                original_name = f'附件{ole_num}.vsd'
                is_image = False
            elif prog_id == 'PBrush':
                bmp_start = ole_data.find(b'BM')
                if bmp_start != -1:
                    bmp_data = ole_data[bmp_start:]
                    try:
                        img = Image.open(io.BytesIO(bmp_data))
                        png_buffer = io.BytesIO()
                        img.save(png_buffer, format='PNG')
                        embedded_data = png_buffer.getvalue()
                        original_name = f'图片{ole_num}.png'
                        is_image = True
                    except:
                        continue
                else:
                    continue

            if original_name and embedded_data and image_name:
                # 清理文件名中的特殊字符，用于保存
                safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', original_name)

                # 保存到对应目录（按需创建）
                if is_image:
                    # 图片统一编号放入images文件夹，使用简短命名
                    images_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                    numbered_name = f"image_{image_idx}.png"
                    save_path = images_dir / numbered_name
                    save_path.write_bytes(embedded_data)
                    print(f"提取图片: {original_name} -> {numbered_name}")
                    ole_info[image_name] = {
                        'original_name': original_name,
                        'saved_name': numbered_name,
                        'is_image': True,
                        'image_idx': image_idx
                    }
                    image_idx += 1
                else:
                    attachments_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                    save_path = attachments_dir / safe_name
                    save_path.write_bytes(embedded_data)
                    print(f"提取附件: {original_name}")
                    ole_info[image_name] = {
                        'original_name': original_name,
                        'saved_name': safe_name,
                        'is_image': False
                    }

    # 处理Excel的md文件生成（在ole_info中检查是否有Excel特殊记录）
    for image_name, info in list(ole_info.items()):
        if 'xlsx_data' in info:  # 这是Excel的特殊记录
            attachments_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
            # 保存xlsx文件
            xlsx_safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', info['xlsx_name'])
            xlsx_path = attachments_dir / xlsx_safe_name
            xlsx_path.write_bytes(info['xlsx_data'])
            print(f"提取附件(xlsx): {info['xlsx_name']}")

            # 保存md文件
            md_safe_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', info['md_name'])
            md_path = attachments_dir / md_safe_name
            md_path.write_text(info['md_content'], encoding='utf-8')
            print(f"生成文本(md): {info['md_name']}")

            # 更新ole_info为标准格式
            ole_info[image_name] = {
                'original_name': info['xlsx_name'],
                'saved_name': xlsx_safe_name,
                'md_name': md_safe_name,
                'is_image': False,
                'has_md': True
            }

    return ole_info, image_idx


def xlsx_to_md(xlsx_path: str, output_dir: str = None) -> str:
    """
    将xlsx/xls文件转换为markdown格式

    Args:
        xlsx_path: xlsx/xls文件路径
        output_dir: 输出目录，默认为xlsx文件所在目录

    Returns:
        生成的md文件路径
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"文件不存在: {xlsx_path}")

    if output_dir is None:
        output_dir = xlsx_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    base_name = xlsx_path.stem
    md_path = output_dir / f"{base_name}.md"

    print(f"处理Excel文件: {xlsx_path}")

    try:
        # 用openpyxl读取数据
        wb = load_workbook(xlsx_path, data_only=True)
        sheet_names = wb.sheetnames

        md_content = ""

        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            md_content += f"## {sheet_name}\n\n"

            # 读取数据转为表格格式
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                # 过滤全空行
                if any(cell is not None for cell in row):
                    cells = [str(cell) if cell is not None else '' for cell in row]
                    md_content += "| " + " | ".join(cells) + " |\n"
                    # 第一行添加表头分隔线
                    if row_count == 0:
                        md_content += "| " + " | ".join(["---"] * len(cells)) + " |\n"
                    row_count += 1

            md_content += "\n"

        wb.close()

        md_path.write_text(md_content, encoding="utf-8")
        print(f"转换完成!")
        print(f"MD文件: {md_path}")
        print(f"包含 {len(sheet_names)} 个工作表")

        return str(md_path)

    except Exception as e:
        raise RuntimeError(f"Excel转换失败: {e}")


def docx_to_md(docx_path: str, output_dir: str = None) -> str:
    """
    将docx文件转换为markdown格式，提取图片和嵌入附件

    Args:
        docx_path: docx文件路径
        output_dir: 输出目录，默认为docx文件所在目录

    Returns:
        生成的md文件路径
    """
    docx_path = Path(docx_path)
    if not docx_path.exists():
        raise FileNotFoundError(f"文件不存在: {docx_path}")

    if output_dir is None:
        output_dir = docx_path.parent
    else:
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

    base_name = docx_path.stem

    # 资源文件夹路径（按需创建）
    assets_dir = output_dir / f"{base_name}_files"
    images_dir = assets_dir / "images"
    attachments_dir = assets_dir / "attachments"

    # 先提取OLE嵌入对象，返回ole信息和下一个图片编号
    ole_info, next_image_idx = extract_ole_objects(docx_path, assets_dir, base_name)

    temp_media_dir = output_dir / "_temp_media"
    md_path = output_dir / f"{base_name}.md"

    cmd = [
        "pandoc",
        str(docx_path),
        "-f", "docx",
        "-t", "gfm",
        "--wrap=none",
        "--extract-media", str(temp_media_dir)
    ]

    print(f"执行命令: {' '.join(cmd)}")
    result = subprocess.run(cmd, capture_output=True, encoding="utf-8")

    if result.returncode != 0:
        print(f"pandoc错误: {result.stderr}")
        raise RuntimeError(f"pandoc转换失败: {result.stderr}")

    content = result.stdout

    extracted_media_path = temp_media_dir / "media"

    # 初始化变量（防止分支未执行时报错）
    image_map = {}
    ole_replacements = {}
    idx = next_image_idx

    if extracted_media_path.exists():
        image_files = sorted(extracted_media_path.iterdir())
        idx = next_image_idx  # 继续OLE图片的编号

        for img_file in image_files:
            if img_file.is_file():
                # 检查是否是OLE对象的预览图
                if img_file.suffix.lower() == '.emf' and img_file.name in ole_info:
                    # OLE对象处理在后面统一进行
                    continue

                # 普通图片，使用简短命名（image_1.png而非长文件名前缀）
                images_dir.mkdir(parents=True, exist_ok=True)  # 按需创建
                new_name = f"image_{idx}{img_file.suffix}"
                new_path = images_dir / new_name
                shutil.copy2(img_file, new_path)
                image_map[img_file.name] = new_name
                print(f"图片重命名: {img_file.name} -> {new_name}")
                idx += 1

        # 处理OLE对象（图片和附件），需要处理同一行多个OLE的情况
        for img_file in image_files:
            if img_file.is_file() and img_file.suffix.lower() == '.emf' and img_file.name in ole_info:
                info = ole_info[img_file.name]
                original_name = info['original_name']
                saved_name = info['saved_name']

                if info['is_image']:
                    # 图片，获取文件名（不带后缀）作为标题
                    name_without_ext = Path(original_name).stem
                    # 先用一个临时标记替换，后续统一处理格式
                    placeholder = f'__OLE_IMG_{img_file.name}__'
                    content = re.sub(
                        rf'!\[([^\]]*)\]\([^)]*{re.escape(img_file.name)}\)',
                        placeholder,
                        content
                    )
                    # 记录占位符对应的实际内容
                    ole_replacements[placeholder] = {
                        'type': 'image',
                        'title': name_without_ext,
                        'path': f'{base_name}_files/images/{saved_name}'
                    }
                    print(f"OLE图片标记: {original_name}")
                else:
                    # 附件
                    placeholder = f'__OLE_ATT_{img_file.name}__'
                    content = re.sub(
                        rf'!\[([^\]]*)\]\([^)]*{re.escape(img_file.name)}\)',
                        placeholder,
                        content
                    )
                    if info.get('has_md'):
                        md_name = info.get('md_name')
                        ole_replacements[placeholder] = {
                            'type': 'attachment',
                            'link': f'[📎 {original_name}]({base_name}_files/attachments/{saved_name}) [📄 纯文本]({base_name}_files/attachments/{md_name})'
                        }
                    else:
                        ole_replacements[placeholder] = {
                            'type': 'attachment',
                            'link': f'[📎 {original_name}]({base_name}_files/attachments/{saved_name})'
                        }
                    print(f"OLE附件标记: {original_name}")

        shutil.rmtree(temp_media_dir)

        # 替换普通图片路径
        for old_name, new_name in image_map.items():
            content = re.sub(
                rf'<img[^>]*src="[^"]*{re.escape(old_name)}"[^>]*>',
                rf'![]({base_name}_files/images/{new_name})',
                content
            )
            content = re.sub(
                rf'!\[([^\]]*)\]\([^)]*{re.escape(old_name)}\)',
                rf'![\1]({base_name}_files/images/{new_name})',
                content
            )

        print(f"已更新md文件中的图片引用")

    # 处理HTML表格中的markdown图片，转换为<img>标签
    # 因为某些markdown渲染器在HTML表格中无法正确解析markdown图片语法

    # 处理 <td>...</td> 中直接包含的markdown图片（单个或多个）
    def replace_md_img_in_td(match):
        td_content = match.group(1)
        # 替换所有markdown图片为img标签
        td_content = re.sub(
            r'!\[\]\(([^)]+)\)',
            r'<img src="\1">',
            td_content
        )
        td_content = re.sub(
            r'!\[([^\]]*)\]\(([^)]+)\)',
            r'<img src="\2" alt="\1">',
            td_content
        )
        return f'<td>{td_content}</td>'

    content = re.sub(
        r'<td[^>]*>(.*?)</td>',
        replace_md_img_in_td,
        content,
        flags=re.DOTALL
    )

    # 处理OLE占位符，逐行处理同一行多个OLE的情况
    lines = content.split('\n')
    result_lines = []
    for line in lines:
        # 检查这一行是否包含OLE占位符
        has_ole = any(placeholder in line for placeholder in ole_replacements)

        if has_ole:
            # 找出这一行所有OLE占位符的顺序
            placeholders_in_line = []
            for placeholder in ole_replacements:
                if placeholder in line:
                    # 找到位置并记录顺序
                    pos = line.find(placeholder)
                    placeholders_in_line.append((pos, placeholder))
            # 按位置排序
            placeholders_in_line.sort(key=lambda x: x[0])

            # 构建新内容，每个OLE都单独一行
            # 先保留行前面的内容
            prefix = line[:placeholders_in_line[0][0]].strip()

            new_parts = []
            if prefix:
                new_parts.append(prefix)

            for pos, placeholder in placeholders_in_line:
                info = ole_replacements[placeholder]
                if info['type'] == 'image':
                    # 图片：标题 + 图片嵌入
                    new_parts.append(f"**{info['title']}**")
                    new_parts.append(f"![]({info['path']})")
                else:
                    # 附件：链接
                    new_parts.append(info['link'])

            # 检查占位符后面是否还有内容
            last_placeholder_end = 0
            for pos, placeholder in placeholders_in_line:
                last_placeholder_end = max(last_placeholder_end, pos + len(placeholder))

            suffix = line[last_placeholder_end:].strip()
            if suffix:
                new_parts.append(suffix)

            # 每个部分单独一行
            line = '\n'.join(new_parts)

        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 处理表格空表头
    lines = content.split('\n')
    new_lines = []
    i = 0
    skip_next = False

    while i < len(lines):
        if skip_next:
            skip_next = False
            i += 1
            continue

        line = lines[i]
        if re.match(r'^\|[\s:|-]+\|$', line.strip()):
            if new_lines:
                prev_line = new_lines[-1]
                if re.match(r'^\|[\s|]+\|$', prev_line.strip()):
                    cells = [c.strip() for c in prev_line.split('|') if c.strip()]
                    if not cells:
                        if i + 1 < len(lines):
                            next_line = lines[i + 1]
                            next_cells = [c.strip() for c in next_line.split('|') if c.strip()]
                            all_bold = all(re.match(r'^\*\*[^*]+\*\*$', c) for c in next_cells)
                            if all_bold and next_cells:
                                new_lines.pop()
                                new_lines.append(next_line)
                                new_lines.append(line)
                                skip_next = True
                                i += 1
                                continue
        new_lines.append(line)
        i += 1

    content = '\n'.join(new_lines)

    # 为标题添加序号
    # 根据层级生成序号: # -> 1., ## -> 1.1, ### -> 1.1.1, #### -> 1.1.1.1
    lines = content.split('\n')
    result_lines = []
    counters = [0, 0, 0, 0, 0]  # 5级标题计数器

    for line in lines:
        # 检测标题层级
        heading_match = re.match(r'^(#{1,6})\s+(.*)$', line)
        if heading_match:
            level = len(heading_match.group(1))
            title_text = heading_match.group(2)

            # 更新计数器
            counters[level - 1] += 1
            # 重置下级计数器
            for i in range(level, 5):
                counters[i] = 0

            # 生成序号（中式格式：1、1.1、1.1.1）
            nums = [str(counters[j]) for j in range(level) if counters[j] > 0]
            if nums:
                number = '.'.join(nums) + ' '
                # 如果标题已有序号开头，不重复添加
                if not re.match(r'^\d+', title_text):
                    line = heading_match.group(1) + ' ' + number + title_text

        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 处理列表中的HTML注释分隔符 <!-- -->
    # pandoc在父列表项后用<!-- -->分隔，子列表项需要缩进
    lines = content.split('\n')
    result_lines = []

    i = 0
    indent_next_items = False  # 标记后续列表项是否需要缩进

    while i < len(lines):
        line = lines[i]
        stripped = line.strip()

        # 处理HTML注释行
        if stripped == '<!-- -->':
            # 回溯找最近的非空列表项
            for j in range(len(result_lines) - 1, max(-1, len(result_lines) - 5), -1):
                prev_stripped = result_lines[j].strip()
                if prev_stripped == '':
                    continue
                if prev_stripped.startswith('-') or re.match(r'^\d+\.', prev_stripped):
                    indent_next_items = True
                    break
                else:
                    break
            i += 1
            continue

        # 处理需要缩进的内容（只处理列表项，附件链接不缩进并去掉前导空格）
        if stripped.startswith('[📎'):
            # 附件链接行，去掉前导空格
            line = stripped
        elif stripped.startswith('-') or re.match(r'^\d+\.', stripped):
            if indent_next_items:
                # 添加缩进
                existing_indent = len(line) - len(line.lstrip())
                line = ' ' * (existing_indent + 4) + line.lstrip()
        else:
            # 其他非空行，恢复缩进标记
            if stripped:
                indent_next_items = False

        result_lines.append(line)
        i += 1

    content = '\n'.join(result_lines)

    # 处理同一行多个附件链接，换行显示
    lines = content.split('\n')
    result_lines = []
    for line in lines:
        # 统计这一行中附件链接的数量（以[📎开头的）
        # 找出所有[📎 xxx](path)模式
        attachment_links = re.findall(r'\[📎[^\]]*\]\([^)]+\)', line)
        if len(attachment_links) > 1:
            # 多个链接，每个链接换行显示
            # 先保留行前面的非链接内容
            prefix_match = re.match(r'^([^[]*)', line)
            prefix = prefix_match.group(1) if prefix_match else ''
            # 构建新行：前缀 + 每个链接换行
            new_content = prefix.rstrip()
            for link in attachment_links:
                # 检查链接后是否有纯文本链接
                plain_text_pattern = rf'{re.escape(link)}\s*\[📄 纯文本\]\([^)]+\)'
                plain_match = re.search(plain_text_pattern, line)
                if plain_match:
                    new_content += '\n' + plain_match.group(0)
                else:
                    new_content += '\n' + link
            line = new_content
        result_lines.append(line)

    content = '\n'.join(result_lines)

    # 更新目录链接格式（去掉页码部分）
    lines = content.split('\n')
    result_lines = []

    for line in lines:
        # 处理目录链接：[N 标题 [页码](#锚点)](#锚点) -> [N 标题](#锚点)
        toc_match = re.match(r'^\[([\d\.\s]+)([^\[]+)\s+\[[^\]]+\]\(#([^)]+)\)\]\(#([^)]+)\)', line)
        if toc_match:
            number_part = toc_match.group(1).strip()
            title_part = toc_match.group(2).strip()
            # 生成锚点：序号-标题，点号用-连接
            anchor_number = re.sub(r'\.', '-', number_part)
            anchor = f'{anchor_number}-{title_part}'.lower()
            # 保留锚点链接（Markdown标题会自动生成锚点）
            line = f'[{number_part} {title_part}](#{anchor})'

        result_lines.append(line)

    content = '\n'.join(result_lines)

    md_path.write_text(content, encoding="utf-8")

    print(f"\n转换完成!")
    print(f"MD文件: {md_path}")

    # 只在存在资源目录时才显示
    has_images = images_dir.exists() and any(images_dir.iterdir())
    has_attachments = attachments_dir.exists() and any(attachments_dir.iterdir())

    if has_images or has_attachments:
        print(f"资源目录: {assets_dir}")
        if has_images:
            print(f"  - 图片: {images_dir}")
        if has_attachments:
            print(f"  - 附件: {attachments_dir}")

    return str(md_path)


def main():
    import sys
    import argparse

    parser = argparse.ArgumentParser(
        description='将文档转换为AI可读的Markdown格式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
支持格式:
  - docx: Word文档，提取图片和OLE附件
  - xlsx/xls: Excel文件，转换为markdown表格

输出结构:
  docx: {filename}.md + {filename}_files/images/ + {filename}_files/attachments/
  xlsx: {filename}.md (包含所有工作表)

示例:
  python docx2md.py document.docx
  python docx2md.py spreadsheet.xlsx
  python docx2md.py document.docx -o ./output
        '''
    )
    parser.add_argument('input_file', help='要转换的文件路径 (docx/xlsx/xls)')
    parser.add_argument('-o', '--output', help='输出目录（默认为输入文件所在目录）')

    args = parser.parse_args()

    input_path = Path(args.input_file)
    suffix = input_path.suffix.lower()

    try:
        if suffix in ['.xlsx', '.xls']:
            md_path = xlsx_to_md(args.input_file, args.output)
        elif suffix == '.docx':
            md_path = docx_to_md(args.input_file, args.output)
        else:
            print(f"不支持的文件格式: {suffix}")
            print("支持的格式: .docx, .xlsx, .xls")
            sys.exit(1)

        print(f"\n转换成功!")
        print(f"MD文件: {md_path}")
    except FileNotFoundError as e:
        print(f"文件不存在: {e}")
        sys.exit(1)
    except RuntimeError as e:
        print(f"转换失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
